<<<<<<< .merge_file_a12004
#import numpy as n
=======
import numpy as n
>>>>>>> .merge_file_a12956
from openpyxl import *
import datetime
from datetime import timedelta

def mean(array):
    avg = 0
    for i in range(len(array)):
        #print((array[i]))
        if type(array[i]) is not None:
            avg = avg + array[i]
    avg = avg/len(array)
    return avg

def getRCMPLcount(ws, numRows):
    #RCMPLcount = []
<<<<<<< .merge_file_a12004
    for i in range(numRows-1, -1, -1):
=======
    for i in range(numRows-1, 0-1, -1):
>>>>>>> .merge_file_a12956
        RCMPLcount[(numRows-1)-i] = ws.cell(row = i+1, column = 3).value
    #print("RCMPLcount = " ,RCMPLcount)
    return RCMPLcount

def average(ws, row):
    for j in range(row,0,-1):
        date = ws.cell(row = j, column = 1).value
        k = j
        sumRCMPL = 0
        sumUnblocked = 0
        hours = 0
        avgRCMPL = 0
        avgUnblocked = 0
        #print(k)
        while ws.cell(row = k, column = 1).value == date:
            sumRCMPL = sumRCMPL + ws.cell(row = k, column = 3).value
            #print(ws.cell(row = k, column = 4).value)
            sumUnblocked = sumUnblocked + ws.cell(row = k, column = 4).value
            #print(sumUnblocked)
            hours = hours + 1
            k = k-1
            if k==0:
                k = 1
                break
        #print("------")
        j = k
        avgRCMPL = sumRCMPL / hours
        avgUnblocked = sumUnblocked / hours
        #print(avgUnblocked)
        return j, avgRCMPL, avgUnblocked

def RCMPLlevelList(ws, row, RCMPLseasonal):
    RCMPLlevel[row] = rLevel*(RCMPLcount[row-1]/seasonal(ws, row-1, RCMPLseasonal)) + (1-rLevel)*(RCMPLlevel[row-1] + RCMPLtrend[row-1])
    RCMPLlvl = rLevel*(RCMPLcount[row-1]/seasonal(ws, row-1, RCMPLseasonal)) + (1-rLevel)*(RCMPLlevel[row-1] + RCMPLtrend[row-1])
    #print(row, " | ", RCMPLlvl, " | ",RCMPLcount[row-1], " | ",seasonal(ws, row-1, RCMPLseasonal), " | ",RCMPLlevel[row-1], " | ",RCMPLtrend[row-1])
    return RCMPLlvl

def RCMPLtrendList(ws, RCMPLcount, RCMPLseasonal):
    nArray = [None]*24
    pArray = [None]*24
    #print(RCMPLcount)
    for i in range(0,24):
        nArray[i] = RCMPLcount[i+24]
        pArray[i] = RCMPLcount[i]
    
<<<<<<< .merge_file_a12004
    RCMPLtrend[0] = (mean(nArray) - mean(pArray))/24
=======
    RCMPLtrend[0] = (n.mean(nArray) - n.mean(pArray))/24
>>>>>>> .merge_file_a12956
    for i in range(1,numRows+1):
        #print(RCMPLlevel[i-1])
        RCMPLtrend[i] = rTrend*(RCMPLlevelList(ws, i, RCMPLseasonal) - RCMPLlevel[i-1]) + (1-rTrend)*RCMPLtrend[i-1]
    return RCMPLtrend

def hrVar(ws, rcmplAvg):
    #hrVarRCMPL = hrVarUnblocked = []
    #print(numRows, numDays)
    #print("Lenghts : ", len(hrVarRCMPL), numRows)
    a = 1
    for i in range(0,numRows):
        for j in range(numDays):
            if(ws.cell(row = numRows-i, column = 1).value == rcmplAvg[j]['date']):
                hrVarRCMPL[i] = RCMPLcount[i]/rcmplAvg[j]['avg']
                a+=1
    #hrVarRCMPL[119]
    #print(a)
    return hrVarRCMPL

def seasonal(ws, row, RCMPLseasonal):
    seasonalInitialRCMPL = ["23:30","22:30","21:30","20:30","19:30","18:30","17:30","16:30","15:30","14:30","13:30","12:30","11:30","10:30",
                            "09:30","08:30","07:30","06:30","05:30","04:30","03:30","02:30","01:30","00:30"]
    
    avg = 0
    count = 0
    #print(len(RCMPLcountMod))
    if row < 24:
        for i in range(len(hrVarRCMPL)):
            #print(time, RCMPLcountMod[i]['time'])
            if seasonalInitialRCMPL[row] == RCMPLcountMod[i]['time']:
                #print(avg)
                avg = avg + hrVarRCMPL[i]
                count += 1
        RCMPLseasonal[row] = avg/count
        return RCMPLseasonal[row]
            
    if row >= 24:
        
        RCMPLseasonal[row] = rSeasonal*(RCMPLcount[row-24]/RCMPLlevel[row-23]) + (1-rSeasonal)*(RCMPLseasonal[row-24])
        #print(RCMPLcount[row-24]," | ", RCMPLlevel[row-23]," | ", rSeasonal ," | ",RCMPLseasonal[row]," | ",row)
        return RCMPLseasonal[row]
       

def forecastSeasonal(RCMPLcountMod):
    timeInitial = ["23:30","22:30","21:30","20:30","19:30","18:30","17:30","16:30","15:30","14:30","13:30","12:30","11:30","10:30",
                   "09:30","08:30","07:30","06:30","05:30","04:30","03:30","02:30","01:30","00:30"]
    avg = 0
    count = 0
    RCMPLseasonalForecast  = [dict() for x in range(24 + numRows)]
    row = 0
    for row in range(0,24):
        for i in range(len(RCMPLcountMod)):
            if timeInitial[row] == RCMPLcountMod[i]['time']:
                avg = avg + RCMPLcountMod[i]['hrVar']
                count += 1
        RCMPLseasonalForecast[row] = {'time' : timeInitial[row], 'seasonal' : avg/count}
        avg = count = 0

    #print(len(RCMPLseasonalForecast), numRows)
    if row >= 23:
        for row in range(24, numRows+24):
            #print(row, RCMPLcountMod[row-24]['RCMPL'], RCMPLlevel[row-23],RCMPLseasonalForecast[row-24]['seasonal'])
            RCMPLseasonalForecast[row] = {'seasonal' : rSeasonal*(RCMPLcountMod[row-24]['RCMPL']/RCMPLlevel[row-23]) +
                                          (1-rSeasonal)*RCMPLseasonalForecast[row-24]['seasonal'], 'time' : RCMPLcountMod[row-24]['time']}
    #print(RCMPLlevel)    
    return RCMPLseasonalForecast

def RCMPLtoMod(ws, RCMPLcount, hrVarRCMPL):
    RCMPLcountMod = [dict() for x in range(0,numRows)]
    #print(numRows)
    for i in range(0, numRows):
        RCMPLcountMod[i] = {'time' : ws.cell(row = i+1, column = 2).value, 'RCMPL' : RCMPLcount[i], 'hrVar' : hrVarRCMPL[i]}
    #print(RCMPLcountMod)
    return RCMPLcountMod

def forecast(RCMPLseasonalForecast, RCMPLtrend, RCMPLlevel):
    #print(RCMPLlevel[-1])
    forecastRCMPL = [None]*24
    for i in range(0,24):
        #print(RCMPLlevel[-1])
        forecastRCMPL[i] = (RCMPLlevel[-1] + RCMPLtrend[-1] + i+1)*RCMPLseasonalForecast[len(RCMPLseasonalForecast)-24+i]['seasonal']
        #print(RCMPLlevel[-1], RCMPLtrend[-1], i, RCMPLseasonalForecast[len(RCMPLseasonalForecast)-25+i]['seasonal'])
    #print("\nThis is forecast \n\n",forecastRCMPL)
    return forecastRCMPL

def meanAPE(rcmplAPE):
<<<<<<< .merge_file_a12004
    return mean(rcmplAPE)
=======
    return n.mean(rcmplAPE)
>>>>>>> .merge_file_a12956
    
def RCMPLmain():
    global RCMPLlevel, RCMPLtrend, RCMPLavg, RCMPLhrVar, RCMPLseasonal, RCMPLcount                                                      #Global input vectors
    global unblockedLevel, unblockedTrend, unblockedAvg, unblockedVar, unblockedSeasonal, unblockedCount
    global RCMPLforecast, rcmplAPE
    global unblockedForecast, unblockedAPE                                                                                              #Global output vectors
    global rTrend, rLevel, rSeasonal, uTrend, uLevel, uSeasonal
    global hrVarRCMPL, RCMPLcountMod

    global numRows, numDays

    RCMPLlevel=RCMPLtrend=RCMPLavg=RCMPLhrVar=RCMPLseasonal=RCMPLcount = []
    unblockedLevel=unblockedTrend=unblockedAvg=unblockedVar=unblockedSeasonal=unblockedCount = []

    #rTrend=rLevel=rSeasonal = 0.50
    rTrend = 0.55
    rLevel = 0.76
    rSeasonal = 0.102
    
    wb = load_workbook("resources/copyAvgSort.xlsx", data_only = True)
    ws = wb.get_sheet_by_name('Sort')
    actForecastRCMPL = [264.75,326.5,307.5833333,262.4166667,249.3333333,235.0833333,250.6666667,276,272.75,280.4166667,283.9166667,261.0833333,258.3333333,235.25,189.5833333,
                        126.75,75.16666667,34.75,19.83333333,26.08333333,44.5,92.83333333,126.3333333,225.8333333]
    
<<<<<<< .merge_file_a12004
    numRows = 192-24
=======
    numRows = ws.get_highest_row()
>>>>>>> .merge_file_a12956
    RCMPLcount = [None] * numRows
    RCMPLlevel = [None] * (numRows+1)
    RCMPLtrend = [None] * (numRows+1)
    rcmplAPE = [None]*24
<<<<<<< .merge_file_a12004
    print "RCMPL Forecast module" , numRows, ws.cell(row=2,column=3).value
=======
    
>>>>>>> .merge_file_a12956
    RCMPLcount = getRCMPLcount(ws, numRows)
    
    #print(RCMPLcount1)
    RCMPLlevel[0] = RCMPLcount[0]
    RCMPLseasonal = [None]*(numRows + 24)
    hrVarRCMPL = [None]*numRows
    row = numRows
    j = 0
    numDays = 0
    
    for i in range(1,numRows):
        if ws.cell(row = i, column = 2).value == "23:30":
            numDays = numDays+1
<<<<<<< .merge_file_a12004
    numDays = numDays + 3
=======
>>>>>>> .merge_file_a12956

    rcmplAvg = [dict() for x in range(numDays+1)]
    unblockedAvg = [dict() for x in range(numDays+1)]

    for i in range(numDays,-1,-1):
        if row >= 1:
            row, avgRCMPL, avgUnblocked = average(ws, row)
            rcmplAvg[j] = {'date' : ws.cell(row = row+1, column = 1).value, 'avg' : avgRCMPL}
            unblockedAvg[j] = {'date' : ws.cell(row = row+1, column = 1).value, 'avg' : avgUnblocked}
            j = j+1
    #print(rcmplAvg)
    time = [0]*24
    for i in range(1,25):
        time[i-1] = ws.cell(row = i, column = 2).value
    #print(time)
    forecastDate = rcmplAvg[-1]['date'] + timedelta(days=1)
    #print(date)
    #hrVarRCMPL = []
    #hrVarUnblocked = []

    hrVarRCMPL = hrVar(ws, rcmplAvg)
    #print(hrVarRCMPL)
    RCMPLcountMod = RCMPLtoMod(ws, RCMPLcount, hrVarRCMPL)
    RCMPLtrend = RCMPLtrendList(ws, RCMPLcount, RCMPLseasonal)    
    #print(RCMPLcountMod)
    #RCMPLseasonal1 = seasonal1()
    RCMPLseasonalForecast = forecastSeasonal(RCMPLcountMod)
    forecastRCMPL = forecast(RCMPLseasonalForecast, RCMPLtrend, RCMPLlevel)
    #d1 = datetime.datetime.strptime(forecastDate.strftime('%d/%m/%Y')+ " "+str(time[0]),'%d/%m/%Y %H:%M')
    for i in range(24):
        rcmplAPE[i] = abs((forecastRCMPL[i]-actForecastRCMPL[i])/actForecastRCMPL[i])
    avg = meanAPE(rcmplAPE)
    #print(avg)
    return forecastRCMPL, forecastDate, time
<<<<<<< .merge_file_a12004

if __name__ == "__main__":    
    forecastRCMPL, forecastDate, forecastTime = RCMPLmain()
    print(forecastRCMPL)
=======
    
forecastRCMPL, forecastDate, forecastTime = RCMPLmain()
print(forecastDate)
>>>>>>> .merge_file_a12956
