from openpyxl import *
import datetime, math, sys, time, os, gc
from MetricsAvgSort import sortMetric, averageMetrics

global toolbar_width

toolbar_width = 40

<<<<<<< .merge_file_a04856
def readInVariable():
    gc.disable()
    print("In readInVariable() function")
    wb = load_workbook('resources/dbMetrics.xlsx', use_iterators = True, read_only = True)
    ws = wb.get_sheet_by_name("metrics")
    numRows = ws.get_highest_row()
    #print(numRows)
    date = []
    time = []
    RCMPL = []
    Unblocked = []
    timeInitial = datetime.datetime.now()
    #print(numRow)
    #print(ws.iter_rows('A2:S'+str(numRows)))
    ws_iter = tuple(ws.iter_rows('A2:D'+str(numRows)))
    #print("11111")
    #print(type(ws_iter))
    i = 0
    j= 1
    for row in ws_iter:
        #if(i%500 == 0):
            #print(i, datetime.datetime.now()-timeInitial)
        for cell in row:
            if j == 1:
                date.append(cell.value)
            elif j == 2:
                time.append(cell.value)
            elif j == 3:
                RCMPL.append(cell.value)
            elif j == 4:
                Unblocked.append(cell.value)
            j = j+1
        j = 1
    print("Length of date ",len(date), len(RCMPL))
##    for i in range(1,numRow):
##        if(i%500 == 0):
##            print(i, datetime.datetime.now()-timeInitial)
##        date[i-1] = ws.cell(row = i, column = 1).value
##        #print(ws.cell(row = 1, column = 1).value)
##        time[i-1] = ws.cell(row = i, column = 2).value
##        RCMPL[i-1] = ws.cell(row = i, column = 3).value
##        Unblocked[i-1] = ws.cell(row = i, column = 4).value
=======
def readInVariable(numRow):
    gc.disable()
    print("In readInVariable() function")
    wb = load_workbook('resources/dbMetrics.xlsx', use_iterators = False, read_only = True)
    ws = wb.get_sheet_by_name("metrics")
    numRows = ws.get_highest_row() - 1
    date = [None]*numRow
    time = [None]*numRow
    RCMPL = [None]*numRow
    Unblocked = [None]*numRow
    timeInitial = datetime.datetime.now()
    print(numRow)
    for i in range(1,numRow):
        if(i%500 == 0):
            #print(i, datetime.datetime.now()-timeInitial)
        date[i-1] = ws.cell(row = i, column = 1).value
        time[i-1] = ws.cell(row = i, column = 2).value
        RCMPL[i-1] = ws.cell(row = i, column = 3).value
        Unblocked[i-1] = ws.cell(row = i, column = 4).value
>>>>>>> .merge_file_a12572
    timeFinal = datetime.datetime.now()
    print("Net time to convert metrics into arrays = {0}".format(timeFinal - timeInitial))
    return date, time, RCMPL, Unblocked

def readInAvgMetrics():
    print("In readInAvgMetrics() function")
    wb = load_workbook('resources/avgSort.xlsx', use_iterators = True, data_only = True, read_only = True)
    ws = wb.get_sheet_by_name('Avg')
    timeInitial = datetime.datetime.now()
    numRows = ws.get_highest_row()
    date = time = RCMPL = Unblocked = RCMPLavg = UnblockedAvg = []
    for i in range(1,numRows):
        date.append(ws.cell(row = i, column = 1).value)
        time.append(ws.cell(row = i, column = 2).value)
        RCMPL.append(ws.cell(row = i, column = 3).value)
        Unblocked.append(ws.cell(row = i, column = 4).value)
        RCMPLavg.append(ws.cell(row = i, column = 5).value)
        UnblockedAvg.append(ws.cell(row = i, column = 6).value)
    timeFinal = datetime.datetime.now()
    print("Net time to convert metrics from avgSort.xlsx into arrays = {0}".format(timeFinal - timeInitial))
    return date, time, RCMPL, Unblocked, RCMPLavg, UnblockedAvg

def avgMetric(initial, final, row, i, bar, wbAvg, wsAvg, ws, wsName, wb, date, time, RCMPL, Unblocked):
<<<<<<< .merge_file_a04856
    dateIni = datetime.datetime.now()
    ii = 0
    #print(time[final][0:2],time[initial][0:2])
##    print(initial,final)
    while(final < len(date)):
        #print(final)
##        if ii==4907:
##            print("w",time[final][0:2] , time[initial][0:2] , time[final + 1][0:2], initial, final,date[initial],date[final],date[final+1],time[final+1])
        if (time[final][0:2] == time[initial][0:2]) & (time[final + 1][0:2] != time[initial][0:2]):                                                     #Same day
            initial, final, average, wbAvg, ii = averageMetrics(wsName,wb, initial, final, wbAvg, wsAvg, date, time, RCMPL, Unblocked, ii)
##            if ii==4907:
##                print("t")
##            if ii>8000:
##                print("ii from true",ii)
##            ii +=1
##            if ii>8000:
##                print("chk",ii)
            #print(initial, final)
##            if ii % 1000==0:
##                print("Final from avgMetrics - True ",final, ii)
##            if ii>8000:
##                print(ii,initial,final)
            bar = bar + 1
        else:
##            if (time[final][0:2] != time[initial][0:2]):
##                print(initial,final,"hh")
##                if ii==4907:
##                    print("f")
                #print("ab")
                #i = final
                j = 1
                while(j!=0):
#                    print("abc")
                    if(time[final][0:2] == time[initial][0:2]):
                        #print(initial,final)
                        initial, final, average, wbAvg, ii = averageMetrics(wsName,wb, initial, final, wbAvg, wsAvg, date, time, RCMPL, Unblocked, ii)
#                        print(initial,final,"Returned")
##                        if ii>8000:
##                            print("ii from false",ii)
                        ii +=1
                        j = 0
#                        if ii % 1000==0:
#                            print("Final from avgMetrics - True ",final, ii)
##                        if ii>8000:
##                            print(ii,initial,final)
=======
    ii = 0
    while(final <= 5999):
        if (time[final][0:2] == time[initial][0:2]) & (time[final + 1][0:2] != time[initial][0:2]):                                                     #Same day
            initial, final, average, wbAvg, ii = averageMetrics(wsName,wb, initial, final, wbAvg, wsAvg, date, time, RCMPL, Unblocked, ii)
            ii +=1
            #print(initial, final)
            bar = bar + 1
        else:
            if (time[final][0:2] != time[initial][0:2]):
                #i = final
                j = 1
                while(j!=0):
                    if(time[final][0:2] == time[initial][0:2]):
                        initial, final, average, wbAvg, ii = averageMetrics(wsName,wb, initial, final, wbAvg, wsAvg, date, time, RCMPL, Unblocked, ii)
                        ii +=1
                        j = 0
>>>>>>> .merge_file_a12572
                        break
                    final= final-1
                    #j+=1
                    bar = bar + 1
        row = row + 1
        if bar%30 == 0:
            sys.stdout.flush()
            bar = 0
    sys.stdout.write("\nAveraging Metrics Complete")
<<<<<<< .merge_file_a04856
    wbAvg = wbAvg.save("resources/avgSort.xlsx")
    print("Saved. Total time :",datetime.datetime.now()-dateIni)
    #return wbAvg

def sortMetrics(initial, final, row, i, bar, wbSort, wsSort, ws1Sort, ws, time):
    ii = 0
    while(final<= len(time)):
        if final%10000 == 0:
            sys.stdout.write("-"+str(final)+",")
        else:
            #sys.stdout.write("-")
            pass
        '''print((((ws.cell(row = final, column = 2).value[0:2]) == (ws.cell(row = initial, column = 2).value[0:2])) &
            ((ws.cell(row = final+1, column = 2).value[0:2]) != (ws.cell(row = initial, column = 2).value[0:2]))))'''
        if (time[final][0:2] == time[initial][0:2]) & (time[final + 1][0:2] != time[initial][0:2]):
            #Same day
=======
    return wbAvg

def sortMetrics(initial, final, row, i, bar, wbSort, wsSort, ws1Sort, ws, time):
    ii = 0
    while(final+11<= 5999):
        sys.stdout.write("-")
        '''print((((ws.cell(row = final, column = 2).value[0:2]) == (ws.cell(row = initial, column = 2).value[0:2])) &
            ((ws.cell(row = final+1, column = 2).value[0:2]) != (ws.cell(row = initial, column = 2).value[0:2]))))'''
        if (time[final][0:2] == time[initial][0:2]) & (time[final + 1][0:2] != time[initial][0:2]):                     #Same day
>>>>>>> .merge_file_a12572
            #print("True ",initial, final)
            wbSort, wsSort, ws1Sort, initial, final, ii = sortMetric(initial, final, wbSort, wsSort, ws1Sort, row, time, ii)
            ii +=1
            bar = bar + 1
        else:
            if (time[final][0:2] != time[initial][0:2]):
                #i = final
                j = 1
                while(j!=0):
                    #print((wsSort.cell(row = i, column = 2).value[0:2] == wsSort.cell(row = initial, column = 2).value[0:2]), i , initial)
                    if(time[final][0:2] == time[initial][0:2]):
                        #print("False ",initial, final)
                        wbSort, wsSort, ws1Sort, initial, final, ii = sortMetric(initial, final, wbSort, wsSort, ws1Sort, row, time, ii)
                        ii+=1
                        j = 0
                        break
                    #i= i-1
                    final -=1
                    #j+=1
                    bar = bar + 1
        row = row + 1
        if bar%30 == 0:
            sys.stdout.flush()
            bar = 0
    sys.stdout.write("Sorting Metrics Complete")
    return wbSort

<<<<<<< .merge_file_a04856
def HWmain():
    #print(__name__)
    try:
        wb = load_workbook('resources/dbMetrics.xlsx', use_iterators = True)
        print(wb.get_sheet_names())
        wsName = raw_input("Enter the sheet you want to select.")
=======
def HWmain():   
    try:
        wb = load_workbook('resources/dbMetrics.xlsx', use_iterators = True)
        print(wb.get_sheet_names())
        wsName = input("Enter the sheet you want to select.")
>>>>>>> .merge_file_a12572
        ws = wb.get_sheet_by_name(wsName)
    except:
        print("Please put dbMetrics.xlsx in correct directory (Resources dir) which contains RCMPL & Unblocked")
        return
    
<<<<<<< .merge_file_a04856
    date, time, RCMPL, Unblocked = readInVariable()
    #print(len(date))
=======
    date, time, RCMPL, Unblocked = readInVariable(6001)
    
>>>>>>> .merge_file_a12572
    wbAvg = Workbook()
    wsAvg = wbAvg.active
    wsAvg.title = "Avg"
    
    #print("Enter the sheet you want to select.")
    #v = ws.cell('B100')
    #print("Element B100 :" , v.value)    
    #print(wbSort.get_sheet_names())
    #ws1Name = input("Enter the sheet you want to select.")   
    
    numRows = ws.get_highest_row()-1

<<<<<<< .merge_file_a04856
    initial = 0
    final = initial + 12
    
    #print(((ws.cell(row = final, column = 2).value[0:2]) == (ws.cell(row = initial, column = 2).value[0:2])) &
            #((ws.cell(row = final+1, column = 2).value[0:2]) != (ws.cell(row = initial, column = 2).value[0:2])))
=======
    initial = 2
    final = initial + 11
    
    print(((ws.cell(row = final, column = 2).value[0:2]) == (ws.cell(row = initial, column = 2).value[0:2])) &
            ((ws.cell(row = final+1, column = 2).value[0:2]) != (ws.cell(row = initial, column = 2).value[0:2])))
>>>>>>> .merge_file_a12572

    print("Working")
    row = 1
    i = 1
    bar = 0
    print(datetime.datetime.now())
<<<<<<< .merge_file_a04856
    avgMetric(initial, final, row, i, bar, wbAvg, wsAvg, ws, wsName, wb, date, time, RCMPL, Unblocked)
    

    wbSort = load_workbook('resources/avgSort.xlsx', read_only = False)
=======
    wbAvg = avgMetric(initial, final, row, i, bar, wbAvg, wsAvg, ws, wsName, wb, date, time, RCMPL, Unblocked)
    wbAvg = wbAvg.save("resources/avgSort.xlsx")

    wbSort = load_workbook('avgSort.xlsx', read_only = False)
>>>>>>> .merge_file_a12572
    wsSort = wbSort.worksheets[0]
    sheets = wbSort.get_sheet_names()
    flag = 1
    for sheet in sheets:
        if sheet == 'Sort':
            ws1Sort = wbSort.get_sheet_by_name('Sort')
            flag = 0
            break
    if flag == 1:                               #If sheet doesn't exist
        ws1Sort = wbSort.create_sheet()
        ws1Sort.title = "Sort"

        
    print("Waiting for Sort Metrics to Sort it all")
    #time.sleep(1)
    print(datetime.datetime.now())
<<<<<<< .merge_file_a04856
    initial = 0
    final = initial + 12
=======
    initial = 1
    final = initial + 11
>>>>>>> .merge_file_a12572
    #date, time, RCMPL, Unblocked, RCMPLavg, UnblockedAvg readInAvgMetrics
    wbSort = sortMetrics(initial, final, row, i, bar, wbSort, wsSort, ws1Sort, ws, time)
    wbSort.save("resources/avgSort.xlsx")
    print(datetime.datetime.now())

<<<<<<< .merge_file_a04856
if __name__ == "__main__":
    HWmain()
=======
HWmain()
>>>>>>> .merge_file_a12572
