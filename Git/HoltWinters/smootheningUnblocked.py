<<<<<<< .merge_file_a11036
#import numpy as n
from openpyxl import *
from math import *
=======
import numpy as n
from openpyxl import *
>>>>>>> .merge_file_a13064

def mean(array):
    avg = 0
    for i in range(len(array)):
<<<<<<< .merge_file_a11036
=======
        #print((array[i]))
>>>>>>> .merge_file_a13064
        if type(array[i]) is not None:
            avg = avg + array[i]
    avg = avg/len(array)
    return avg

def getUnblockedCount(ws, numRows):
<<<<<<< .merge_file_a11036
    for i in range(numRows-1, 0-1, -1):
        unblockedCount[(numRows-1)-i] = ws.cell(row = i+1, column = 4).value
=======
    #unblockedCount = []
    for i in range(numRows-1, 0-1, -1):
        unblockedCount[(numRows-1)-i] = ws.cell(row = i+1, column = 4).value
    #print(unblockedCount[119])
>>>>>>> .merge_file_a13064
    return unblockedCount

def average(ws, row):
    for j in range(row,0,-1):
        date = ws.cell(row = j, column = 1).value
        k = j
        sumRCMPL = sumUnblocked = hours = 0
        sumRCMPL = avgUnblocked = 0
        while ws.cell(row = k, column = 1).value == date:
            sumRCMPL = sumRCMPL + ws.cell(row = k, column = 3).value
            sumUnblocked = sumUnblocked + ws.cell(row = k, column = 4).value
            hours = hours + 1
            k = k-1
            if k==0:
                k=1
                break
        j = k
        avgRCMPL = sumRCMPL / hours
        avgUnblocked = sumUnblocked / hours
<<<<<<< .merge_file_a11036
=======
        #print(avgUnblocked)
>>>>>>> .merge_file_a13064
        return j, avgRCMPL, avgUnblocked

def unblockedLevelList(ws, row, unblockedSeasonal):
    unblockedLevel[row] = rLevel*(unblockedCount[row-1]/seasonal(ws, row-1, unblockedSeasonal)) + (1-rLevel)*(unblockedLevel[row-1] + unblockedTrend[row-1])
    unblockedlvl = rLevel*(unblockedCount[row-1]/seasonal(ws, row-1, unblockedSeasonal)) + (1-rLevel)*(unblockedLevel[row-1] + unblockedTrend[row-1])
<<<<<<< .merge_file_a11036
=======
    #print(row, " | ", unblockedlvl, " | ",unblockedCount[row-1], " | ",seasonal(ws, row-1, unblockedSeasonal), " | ",unblockedLevel[row-1], " | ",unblockedTrend[row-1])
>>>>>>> .merge_file_a13064
    return unblockedlvl

def unblockedTrendList(ws, unblockedCount, unblockedSeasonal):
    nArray = [None]*24
    pArray = [None]*24
<<<<<<< .merge_file_a11036
=======
    #print(unblockedCount)
>>>>>>> .merge_file_a13064
    for i in range(0,24):
        nArray[i] = unblockedCount[i+24]
        pArray[i] = unblockedCount[i]
    
<<<<<<< .merge_file_a11036
    unblockedTrend[0] = (mean(nArray) - mean(pArray))/24
=======
    unblockedTrend[0] = (n.mean(nArray) - n.mean(pArray))/24
>>>>>>> .merge_file_a13064
    for i in range(1,numRows+1):
        unblockedTrend[i] = rTrend*(unblockedLevelList(ws, i, unblockedSeasonal) - unblockedLevel[i-1]) + (1-rTrend)*unblockedTrend[i-1]
    return unblockedTrend

def hrVar(ws, unblockedAvg):
<<<<<<< .merge_file_a11036
    #hrVarUnblocked = []
    #print("Lenghts : ", len(unblockedAvg), numRows)
    a = 1
    for i in range(0,numRows):
        for j in range(numDays):
            if(ws.cell(row = numRows-i, column = 1).value == unblockedAvg[j]['date']):
                hrVarUnblocked[i] = unblockedCount[i]/unblockedAvg[j]['avg']
                a+=1
    #print "length hr var" , len(hrVarUnblocked)
=======
    #hrVarunblocked = hrVarUnblocked = []
   # print("Lenghts : ", len(hrVarunblocked), numRows)
    a = 1
    #print(unblockedAvg)
    for i in range(0,numRows):
        for j in range(numDays):
            #print(ws.cell(row = numRows-i, column = 1).value , unblockedAvg[j]['date'],i,j)
            if(ws.cell(row = numRows-i, column = 1).value == unblockedAvg[j]['date']):
                #print(unblockedAvg[5]['avg'])
                hrVarUnblocked[i] = unblockedCount[i]/unblockedAvg[j]['avg']
                a+=1
                #print(i,hrVarUnblocked[i])
    #hrVarunblocked[119]
    #print(hrVarUnblocked[9])
>>>>>>> .merge_file_a13064
    return hrVarUnblocked

def seasonal(ws, row, unblockedSeasonal):
    seasonalInitialunblocked = ["23:30","22:30","21:30","20:30","19:30","18:30","17:30","16:30","15:30","14:30","13:30","12:30","11:30","10:30",
                            "09:30","08:30","07:30","06:30","05:30","04:30","03:30","02:30","01:30","00:30"]
    
    avg = 0
    count = 0
<<<<<<< .merge_file_a11036
    if row < 24:
        for i in range(len(hrVarUnblocked)):
            if seasonalInitialunblocked[row] == unblockedCountMod[i]['time']:
=======
    #print(len(unblockedCountMod))
    if row < 24:
        for i in range(len(hrVarUnblocked)):
            #print(time, unblockedCountMod[i]['time'])
            if seasonalInitialunblocked[row] == unblockedCountMod[i]['time']:
                #print(avg)
>>>>>>> .merge_file_a13064
                avg = avg + hrVarUnblocked[i]
                count += 1
        unblockedSeasonal[row] = avg/count
        return unblockedSeasonal[row]
            
    if row >= 24:
        
        unblockedSeasonal[row] = rSeasonal*(unblockedCount[row-24]/unblockedLevel[row-23]) + (1-rSeasonal)*(unblockedSeasonal[row-24])
<<<<<<< .merge_file_a11036
=======
        #print(unblockedCount[row-24]," | ", unblockedLevel[row-23]," | ", rSeasonal ," | ",unblockedSeasonal[row]," | ",row)
>>>>>>> .merge_file_a13064
        return unblockedSeasonal[row]
       

def forecastSeasonal(unblockedCountMod):
    timeInitial = ["23:30","22:30","21:30","20:30","19:30","18:30","17:30","16:30","15:30","14:30","13:30","12:30","11:30","10:30",
                   "09:30","08:30","07:30","06:30","05:30","04:30","03:30","02:30","01:30","00:30"]
    avg = 0
    count = 0
    unblockedSeasonalForecast  = [dict() for x in range(24 + numRows)]
    row = 0
    for row in range(0,24):
        for i in range(len(unblockedCountMod)):
            if timeInitial[row] == unblockedCountMod[i]['time']:
                avg = avg + unblockedCountMod[i]['hrVar']
                count += 1
        unblockedSeasonalForecast[row] = {'time' : timeInitial[row], 'seasonal' : avg/count}
        avg = count = 0

<<<<<<< .merge_file_a11036
    if row >= 23:
        for row in range(24, numRows+24):
            unblockedSeasonalForecast[row] = {'seasonal' : rSeasonal*(unblockedCountMod[row-24]['unblocked']/unblockedLevel[row-23]) +
                                          (1-rSeasonal)*unblockedSeasonalForecast[row-24]['seasonal'], 'time' : unblockedCountMod[row-24]['time']} 
=======
    #print(len(unblockedSeasonalForecast), numRows)
    if row >= 23:
        for row in range(24, numRows+24):
            #print(row, unblockedCountMod[row-24]['unblocked'], unblockedLevel[row-23],unblockedSeasonalForecast[row-24]['seasonal'])
            unblockedSeasonalForecast[row] = {'seasonal' : rSeasonal*(unblockedCountMod[row-24]['unblocked']/unblockedLevel[row-23]) +
                                          (1-rSeasonal)*unblockedSeasonalForecast[row-24]['seasonal'], 'time' : unblockedCountMod[row-24]['time']}
    #print(unblockedLevel)    
>>>>>>> .merge_file_a13064
    return unblockedSeasonalForecast

def unblockedtoMod(ws, unblockedCount, hrVarunblocked):
    unblockedCountMod = [dict() for x in range(0,numRows)]
<<<<<<< .merge_file_a11036
=======
    #print(numRows)
>>>>>>> .merge_file_a13064
    for i in range(0, numRows):
        unblockedCountMod[i] = {'time' : ws.cell(row = i+1, column = 2).value, 'unblocked' : unblockedCount[i], 'hrVar' : hrVarunblocked[i]}
    return unblockedCountMod

def forecast(unblockedSeasonalForecast, unblockedTrend, unblockedLevel):
<<<<<<< .merge_file_a11036
    forecastUnblocked = [None]*24
    for i in range(0,24):
        forecastUnblocked[i] = (unblockedLevel[-1] + unblockedTrend[-1] + i+1)*unblockedSeasonalForecast[len(unblockedSeasonalForecast)-24+i]['seasonal']
=======
    #print(unblockedLevel[-1])
    forecastUnblocked = [None]*24
    for i in range(0,24):
        #print(len(unblockedSeasonalForecast)-(25-i))
        forecastUnblocked[i] = (unblockedLevel[-1] + unblockedTrend[-1] + i+1)*unblockedSeasonalForecast[len(unblockedSeasonalForecast)-24+i]['seasonal']
        #print(unblockedLevel[-1], unblockedTrend[-1], i, unblockedSeasonalForecast[len(unblockedSeasonalForecast)-25+i]['seasonal'])
    #print("\nThis is forecast \n\n",forecastunblocked)
>>>>>>> .merge_file_a13064
    return forecastUnblocked
    
def unblockedMain():
    global unblockedLevel, unblockedTrend, unblockedAvg, unblockedVar, unblockedSeasonal, unblockedCount                                                      #Global input vectors
    global unblockedForecast, unblockedAPE                                                                                              #Global output vectors
    global rTrend, rLevel, rSeasonal, uTrend, uLevel, uSeasonal
    global hrVarUnblocked, unblockedCountMod

    global numRows, numDays

    unblockedLevel=unblockedTrend=unblockedAvg=unblockedVar=unblockedSeasonal=unblockedCount = []
    unblockedLevel=unblockedTrend=unblockedAvg=unblockedVar=unblockedSeasonal=unblockedCount = []

<<<<<<< .merge_file_a11036
    rTrend = 0.55
    rLevel = 0.76
    rSeasonal = 0.102

    timeCheck = ["23:30","22:30","21:30","20:30","19:30","18:30","17:30","16:30","15:30","14:30","13:30","12:30","11:30","10:30",
                   "09:30","08:30","07:30","06:30","05:30","04:30","03:30","02:30","01:30","00:30"]
=======
    #rTrend=rLevel=rSeasonal = 0.50
    rTrend = 0.55
    rLevel = 0.76
    rSeasonal = 0.102
>>>>>>> .merge_file_a13064
    
    wb = load_workbook("resources/copyAvgSort.xlsx", data_only = True)
    ws = wb.get_sheet_by_name('Sort')
    
<<<<<<< .merge_file_a11036
    numRows = 192
    unblockedCount = [None] * numRows
    unblockedLevel = [None] * (numRows+1)
    unblockedTrend = [None] * (numRows+1)
    #print "Unblocked Forecast module" , numRows, ws.cell(row=2,column=3).value
    unblockedCount = getUnblockedCount(ws, numRows)
    hrVarUnblocked = [None]*numRows

=======
    numRows = ws.get_highest_row()
    unblockedCount = [None] * numRows
    unblockedLevel = [None] * (numRows+1)
    unblockedTrend = [None] * (numRows+1)
    
    unblockedCount = getUnblockedCount(ws, numRows)
    
    #print(unblockedCount1)
>>>>>>> .merge_file_a13064
    unblockedLevel[0] = unblockedCount[0]
    unblockedSeasonal = [None]*(numRows + 24)
    hrVarUnblocked = [None]*numRows
    row = numRows
    j = 0
    numDays = 0
    
    for i in range(1,numRows):
<<<<<<< .merge_file_a11036
        if ws.cell(row = i, column = 2).value =="23:30":
            numDays = numDays+1
    #print numDays , "= Num days"
    numDays = numDays+3
=======
        if ws.cell(row = i, column = 2).value == "23:30":
            numDays = numDays+1
>>>>>>> .merge_file_a13064

    rcmplAvg = [dict() for x in range(numDays+1)]
    unblockedAvg = [dict() for x in range(numDays+1)]

    for i in range(numDays,-1,-1):
        if row >=1:
            row, avgRCMPL, avgUnblocked = average(ws, row)
            rcmplAvg[j] = {'date' : ws.cell(row = row+1, column = 1).value, 'avg' : avgRCMPL}
            unblockedAvg[j] = {'date' : ws.cell(row = row+1, column = 1).value, 'avg' : avgUnblocked}
            j = j+1
<<<<<<< .merge_file_a11036

    hrVarUnblocked = hrVar(ws, unblockedAvg)
    unblockedCountMod = unblockedtoMod(ws, unblockedCount, hrVarUnblocked)
    unblockedTrend = unblockedTrendList(ws, unblockedCount, unblockedSeasonal)    
    unblockedSeasonalForecast = forecastSeasonal(unblockedCountMod)
    forecastUnblocked = forecast(unblockedSeasonalForecast, unblockedTrend, unblockedLevel)
    return forecastUnblocked

if __name__ == "__main__":
    print unblockedMain()
=======
            
    #hrVarunblocked = []
    #hrVarUnblocked = []

    hrVarUnblocked = hrVar(ws, unblockedAvg)
    #print(rcmplAvg)
    unblockedCountMod = unblockedtoMod(ws, unblockedCount, hrVarUnblocked)
    unblockedTrend = unblockedTrendList(ws, unblockedCount, unblockedSeasonal)    
    #print(unblockedCountMod)
    #unblockedSeasonal1 = seasonal1()
    unblockedSeasonalForecast = forecastSeasonal(unblockedCountMod)
    forecastUnblocked = forecast(unblockedSeasonalForecast, unblockedTrend, unblockedLevel)
    #print(forecastUnblocked)
    return forecastUnblocked
    
unblockedMain() 
>>>>>>> .merge_file_a13064
