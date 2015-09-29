from openpyxl import *
import datetime, math

global timeHour, midCell
timeHour = ["23:30","22:30","21:30","20:30","19:30","18:30","17:30","16:30","15:30","14:30","13:30","12:30","11:30","10:30","09:30","08:30","07:30","06:30","05:30","04:30","03:30","02:30",
            "01:30","00:30"]
midCell = []

def sortMetric(initial, final, wb, ws, ws1, r, time, ii):
    i=1
    #ws1.cell(row = 1, column = 1).value = "Hello"
    for hour in timeHour:
        #print(math.ceil((initial+final)/2), initial, final)
        #print(ws.cell(row = math.ceil((initial+final)/2), column = 2).value[0:2], hour[0:2])
        #print(hour[0:2], ws.cell(row = math.ceil((initial+final)/2), column = 2).value[0:2])
        #x = datetime.datetime.strptime(time[math.ceil((initial+final)/2)], '%H%M%S')
        if (hour[0:2] == time[int(math.ceil((initial+final)/2))][0:2]):
            #print("Iter {0}".format(i))
            #i = i+1
            ws1.cell(row = r, column = 2).value = hour
            #print(ws.cell(row = math.ceil((initial+final)/2), column = 3).value)
            ws1.cell(row = r, column = 1).value = ws.cell(row = midCell[ii], column = 1).value
            ws1.cell(row = r, column = 3).value = ws.cell(row = midCell[ii], column = 5).value
            ws1.cell(row = r, column = 4).value = ws.cell(row = midCell[ii], column = 6).value
    initial = final + 1
    final = initial + 11
    #print("Check ", initial, final)
    return wb, ws, ws1, initial, final, ii


def averageMetrics(ws, wb, initial, final, wb1, ws2, date, time, RCMPL, Unblocked, ii):
##    print(initial, final)
    average = []
    ws1 = wb.get_sheet_by_name(ws)
    numRows = ws1.get_highest_row()-1

    #average = 1
    #testStr = "A"+ str(initial)
    #print(testStr)
    #print("Func : ", initial, final)

    for i in range(initial, final+1):
        ws2.cell(row=i+1, column=1).value = date[i]
        #print(i, ws2.cell(row=i-1, column=1).value, date[i-1])
        ws2.cell(row=i+1, column=2).value = time[i]
        #print(i,ws2.cell(row=i-1, column=2).value, time[i-1])
        ws2.cell(row=i+1, column=3).value = RCMPL[i]
        #print(i, ws2.cell(row=i-1, column=3).value, RCMPL[i-1])
        ws2.cell(row=i+1, column=4).value = Unblocked[i]
        #print(i, ws2.cell(row=i-1, column=4).value, Unblocked[i-1])

    RCMPLstartRange = "C"+str(initial+1)
    i = initial+1
    if final-initial !=12:
        finalR = final+1
    else:
        finalR = final
    RCMPLendRange = "C"+str(finalR)
    rcmplRange = RCMPLstartRange + ":" + RCMPLendRange

    if initial == 0:
        unblockedStartRange = "D"+str(initial+1)
        i = initial+1
    else:
        unblockedStartRange = "D"+str(initial)
        i = initial
    unblockedEndRange = "D"+str(final)
    unblockedRange = unblockedStartRange + ":" + unblockedEndRange
##    if ii>4900:
##        print(ii)
    midCell.append(int(math.ceil((i+final)/2)))
    #print((initialR+finalR)/2)
    ws2["E"+str(int(math.ceil((i+final)/2)))] = "=AVERAGE(Avg!" + rcmplRange + ")"
    #print(ws2["E"+str(midCell[ii])].value, math.ceil((i+final)/2), initial, final)
    ws2["F"+str(int(math.ceil((i+final)/2)))] = "=AVERAGE(Avg!" + unblockedRange + ")" 
    RCMPLcellRange = RCMPLstartRange + ":" + RCMPLendRange
    unblockedCellRange = unblockedStartRange + ":" + unblockedEndRange
    #ws2.garbage_collect()
    #wb1.save("test.xlsx")
##    if ii>8000:
##        print("1-",initial, final)
    average = {'RCMPL': ws2["E"+str(int(math.ceil((i+final)/2)))].value, 'unblocked': ws2["F"+str(int(math.ceil((i+final)/2)))].value}
    #print("call")
    #print("++",initial, final)
    initial = final+1
    final = initial + 12
    #print("+",initial, final)
##    if ii>8000:
##        print("2-",initial, final)
    if ii % 1000==0:
        print("Final from avgerageMetrics ",final, ii)
    return initial, final, average, wb1, ii
